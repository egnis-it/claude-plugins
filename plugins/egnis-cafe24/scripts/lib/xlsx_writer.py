"""
scripts/lib/xlsx_writer.py — CSV/dict rows → xlsx 변환 + 누적 append

설계:
  - 비개발자 사용자 대상 → openpyxl 자동 부트스트랩 (try import / pip install --user)
  - mall_id 별 시트 분리, 헤더 고정 (frozen row 1)
  - 누적 모드(append_rows): 기존 xlsx가 있으면 열어서 행 추가, 없으면 새로 생성
  - 한국어 컬럼명 우선, ID 기반 컬럼은 그대로
  - 부트스트랩 실패 시 RuntimeError → 호출자가 폴백(CSV-only)

ADR 메모:
  - urllib only 정책은 cafe24_client.py 한정.
  - xlsx 출력은 사용자 편의 기능이므로 openpyxl 예외 허용 (auto-install).
"""
from __future__ import annotations

import subprocess
import sys
from pathlib import Path
from typing import Iterable, Sequence


def _ensure_openpyxl():
    """openpyxl import. 없으면 pip install --user 1회 시도."""
    try:
        import openpyxl  # noqa: F401
        return
    except ImportError:
        pass

    print("[xlsx_writer] openpyxl 미설치 — pip install --user 자동 설치 시도...",
          file=sys.stderr)
    try:
        subprocess.run(
            [sys.executable, "-m", "pip", "install", "--user", "--quiet", "openpyxl"],
            check=True,
            timeout=120,
        )
    except (subprocess.CalledProcessError, subprocess.TimeoutExpired) as e:
        raise RuntimeError(
            "openpyxl auto-install 실패. 수동으로 설치 필요:\n"
            f"  {sys.executable} -m pip install --user openpyxl\n"
            f"원인: {e}"
        ) from e

    try:
        import openpyxl  # noqa: F401
    except ImportError as e:
        raise RuntimeError(
            "openpyxl 설치는 완료됐으나 import 실패. PYTHONPATH/user-site 확인 필요."
        ) from e


# 주문 도메인 컬럼 표준 (CSV 헤더와 1:1 대응)
#
# **PII 적재 금지 정책 (2026-05-19 결정):**
# buyer_name, buyer_email, buyer_cellphone 같은 회원 개인정보는 **수집 대상에서 제외**.
# 매출 집계/분석 목적에는 익명화된 주문ID만 있으면 충분하며, PII 적재는 개인정보보호
# 관점에서 위험이 큼. 절대 컬럼 추가 금지.
ORDERS_DETAIL_COLUMNS: list[tuple[str, str]] = [
    # (csv_key, korean_label)
    ("mall_id", "몰ID"),
    ("brand_label", "브랜드"),
    ("order_id", "주문번호"),
    ("order_date", "주문일시"),
    ("payment_date", "결제일시"),
    ("order_status", "주문상태"),
    ("payment_method_name", "결제수단"),
    ("items_count", "상품수량"),
    ("payment_amount", "결제금액"),
    ("actual_payment_amount", "실결제금액"),
    ("shipping_fee", "배송비"),
    ("currency", "통화"),
    ("canceled", "취소여부"),
    ("refunded", "환불여부"),
]

ORDERS_SUMMARY_COLUMNS: list[tuple[str, str]] = [
    ("date", "일자"),
    ("mall_id", "몰ID"),
    ("brand_label", "브랜드"),
    ("order_count", "주문건수"),
    ("payment_count", "결제건수"),
    ("cancel_count", "취소건수"),
    ("cancel_rate", "취소율"),
    ("gmv", "매출액"),
    ("net_revenue", "순매출"),
    ("avg_order_value", "평균주문단가"),
]


def write_rows_to_xlsx(
    xlsx_path: Path,
    sheet_name: str,
    columns: Sequence[tuple[str, str]],
    rows: Iterable[dict],
    *,
    append: bool = False,
) -> int:
    """xlsx 파일에 행을 작성한다.

    Args:
        xlsx_path: 출력 파일 경로
        sheet_name: 시트 이름 (mall_id, "전체" 등)
        columns: [(csv_key, korean_label), ...] 순서대로 컬럼 정의
        rows: dict 형태의 행 iterable
        append: True면 기존 파일을 열어 같은 시트에 행을 추가. False면 덮어씀.

    Returns:
        실제로 작성된 행 수.
    """
    _ensure_openpyxl()
    from openpyxl import Workbook, load_workbook

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    if append and xlsx_path.exists():
        wb = load_workbook(xlsx_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
            ws.append([label for _, label in columns])
            ws.freeze_panes = "A2"
    else:
        wb = Workbook()
        # 기본 시트 제거 후 명시 시트로 시작
        default = wb.active
        wb.remove(default)
        ws = wb.create_sheet(title=sheet_name)
        ws.append([label for _, label in columns])
        ws.freeze_panes = "A2"

    written = 0
    for row in rows:
        ws.append([_serialize_cell(row.get(key)) for key, _ in columns])
        written += 1

    # 컬럼 너비 자동 조정 (간단 휴리스틱: 헤더 길이 + 4)
    for idx, (_, label) in enumerate(columns, start=1):
        col_letter = ws.cell(row=1, column=idx).column_letter
        ws.column_dimensions[col_letter].width = max(10, len(label) * 2 + 4)

    wb.save(xlsx_path)
    return written


def write_multi_sheet_xlsx(
    xlsx_path: Path,
    sheets: dict[str, tuple[Sequence[tuple[str, str]], Iterable[dict]]],
    *,
    append: bool = False,
) -> dict[str, int]:
    """여러 시트를 한 번에 작성/append.

    Args:
        xlsx_path: 출력 파일
        sheets: {sheet_name: (columns, rows), ...}
        append: True면 기존 파일 열어서 시트별 행 추가.

    Returns:
        시트별 작성 행 수.
    """
    _ensure_openpyxl()
    from openpyxl import Workbook, load_workbook

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    if append and xlsx_path.exists():
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()
        default = wb.active
        wb.remove(default)

    counts: dict[str, int] = {}
    for sheet_name, (columns, rows) in sheets.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
            ws.append([label for _, label in columns])
            ws.freeze_panes = "A2"
            for idx, (_, label) in enumerate(columns, start=1):
                col_letter = ws.cell(row=1, column=idx).column_letter
                ws.column_dimensions[col_letter].width = max(10, len(label) * 2 + 4)

        written = 0
        for row in rows:
            ws.append([_serialize_cell(row.get(key)) for key, _ in columns])
            written += 1
        counts[sheet_name] = written

    wb.save(xlsx_path)
    return counts


def _serialize_cell(value):
    """xlsx 셀에 적합한 형태로 변환. dict/list는 JSON-ish 문자열로."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Y" if value else "N"
    if isinstance(value, (int, float, str)):
        return value
    return str(value)
